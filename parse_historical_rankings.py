import os
import re
import sqlite3
import openpyxl
import pypdf

# Configuration
DB_PATH = "elo_ratings.sqlite"
INPUT_DIR = "data_rankings"

# Event name mapping for sheets
EVENT_MAP = {
    "men's singles": "MS",
    "women's singles": "WS",
    "men's doubles": "MD",
    "women's doubles": "WD",
    "mixed doubles": "XD",
    "mens singles": "MS",
    "womens singles": "WS",
    "mens doubles": "MD",
    "womens doubles": "WD",
}

# Layout-aware helpers and state for PDF parsing
KNOWN_NAMES = None

def parse_date_week_from_filename(filename):
    # Filename format: WR_YYYY-MM-DD_Week_X.xlsx/pdf
    match = re.search(r"WR_(\d{4}-\d{2}-\d{2})_Week_(\d+)", filename)
    if match:
        date_str = match.group(1)
        if date_str.startswith("2101-"):
            date_str = date_str.replace("2101-", "2021-")
        elif date_str.startswith("2002-"):
            date_str = date_str.replace("2002-", "2022-")
        return date_str, int(match.group(2))
    return None, None

def parse_xlsx(file_path, rank_date, week):
    print(f"Parsing Excel: {os.path.basename(file_path)}")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    rows_to_insert = []
    
    for sheet in wb.worksheets:
        sheet_name_lower = sheet.title.lower().strip()
        event = EVENT_MAP.get(sheet_name_lower)
        if not event:
            # Fallback check
            for key, val in EVENT_MAP.items():
                if key in sheet_name_lower:
                    event = val
                    break
        
        if not event:
            continue
            
        print(f"  Processing sheet: {sheet.title} -> {event}")
        is_doubles = event in ["MD", "WD", "XD"]
        
        # Scan for header row
        header_row_idx = None
        headers = []
        
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row and len(row) > 0 and str(row[0]).strip().lower() in ["ranking", "rank"]:
                header_row_idx = idx
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                break
                
        if header_row_idx is None:
            print(f"    [!] Warning: Could not find header row in sheet: {sheet.title}")
            continue
            
        # Map headers
        col_map = {}
        for c_idx, h in enumerate(headers):
            h_lower = h.lower()
            if "ranking" in h_lower or "rank" == h_lower:
                col_map["rank"] = c_idx
            elif "bwf id" in h_lower or "player id" in h_lower or "member id" in h_lower:
                if "p1" in h_lower:
                    col_map["p1_id"] = c_idx
                elif "p2" in h_lower:
                    col_map["p2_id"] = c_idx
                else:
                    col_map["p_id"] = c_idx
            elif "last name" in h_lower or "family name" in h_lower:
                if "p1" in h_lower:
                    col_map["p1_last"] = c_idx
                elif "p2" in h_lower:
                    col_map["p2_last"] = c_idx
                else:
                    col_map["p_last"] = c_idx
            elif "first name" in h_lower or "given name" in h_lower:
                if "p1" in h_lower:
                    col_map["p1_first"] = c_idx
                elif "p2" in h_lower:
                    col_map["p2_first"] = c_idx
                else:
                    col_map["p_first"] = c_idx
            elif "name" in h_lower or "player" in h_lower:
                if "p1" in h_lower:
                    col_map["p1_name"] = c_idx
                elif "p2" in h_lower:
                    col_map["p2_name"] = c_idx
                else:
                    col_map["p_name"] = c_idx
            elif "country" in h_lower or "noc" in h_lower or "nation" in h_lower:
                if "p1" in h_lower:
                    col_map["p1_country"] = c_idx
                elif "p2" in h_lower:
                    col_map["p2_country"] = c_idx
                else:
                    col_map["country"] = c_idx
            elif "points" in h_lower or "point" == h_lower:
                col_map["points"] = c_idx
                
        # Parse data rows
        row_count = 0
        for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
            # Check if row is empty or rank is empty
            rank_val = row[col_map.get("rank")] if "rank" in col_map else None
            if rank_val is None or str(rank_val).strip() == "":
                break # Reached the end of data
                
            try:
                rank = int(float(str(rank_val).strip()))
            except ValueError:
                continue # Header or invalid row
                
            points_val = row[col_map.get("points")] if "points" in col_map else 0
            try:
                # Remove separators and parse as float/int
                points = int(float(str(points_val).replace(",", "").strip())) if points_val is not None else 0
            except ValueError:
                points = 0
                
            if not is_doubles:
                p_id_val = row[col_map.get("p_id")] if "p_id" in col_map else None
                if not p_id_val:
                    continue
                p_id = int(float(str(p_id_val).strip()))
                
                # Extract player name
                if "p_name" in col_map and row[col_map["p_name"]]:
                    p_name = str(row[col_map["p_name"]]).strip()
                elif "p_last" in col_map and "p_first" in col_map:
                    last = str(row[col_map["p_last"]]) if row[col_map["p_last"]] else ""
                    first = str(row[col_map["p_first"]]) if row[col_map["p_first"]] else ""
                    p_name = f"{last} {first}".strip()
                else:
                    p_name = f"Player {p_id}"
                    
                country = str(row[col_map["country"]]).strip() if "country" in col_map and row[col_map["country"]] else ""
                
                rows_to_insert.append((rank_date, week, event, rank, p_id, p_name, country, points))
                row_count += 1
            else:
                # Double players
                p1_id_val = row[col_map.get("p1_id")] if "p1_id" in col_map else None
                p2_id_val = row[col_map.get("p2_id")] if "p2_id" in col_map else None
                
                if p1_id_val:
                    p1_id = int(float(str(p1_id_val).strip()))
                    if "p1_name" in col_map and row[col_map["p1_name"]]:
                        p1_name = str(row[col_map["p1_name"]]).strip()
                    elif "p1_last" in col_map and "p1_first" in col_map:
                        last = str(row[col_map["p1_last"]]) if row[col_map["p1_last"]] else ""
                        first = str(row[col_map["p1_first"]]) if row[col_map["p1_first"]] else ""
                        p1_name = f"{last} {first}".strip()
                    else:
                        p1_name = f"Player {p1_id}"
                        
                    p1_country = str(row[col_map["p1_country"]]).strip() if "p1_country" in col_map and row[col_map["p1_country"]] else ""
                    rows_to_insert.append((rank_date, week, event, rank, p1_id, p1_name, p1_country, points))
                    row_count += 1
                    
                if p2_id_val:
                    p2_id = int(float(str(p2_id_val).strip()))
                    if "p2_name" in col_map and row[col_map["p2_name"]]:
                        p2_name = str(row[col_map["p2_name"]]).strip()
                    elif "p2_last" in col_map and "p2_first" in col_map:
                        last = str(row[col_map["p2_last"]]) if row[col_map["p2_last"]] else ""
                        first = str(row[col_map["p2_first"]]) if row[col_map["p2_first"]] else ""
                        p2_name = f"{last} {first}".strip()
                    else:
                        p2_name = f"Player {p2_id}"
                        
                    p2_country = str(row[col_map["p2_country"]]).strip() if "p2_country" in col_map and row[col_map["p2_country"]] else ""
                    rows_to_insert.append((rank_date, week, event, rank, p2_id, p2_name, p2_country, points))
                    row_count += 1
                    
        print(f"    [OK] Parsed {row_count} rows from {event}.")
        
    wb.close()
    return rows_to_insert

def clean_text(text):
    if not text:
        return ""
    return text.replace("\u00a0", " ").replace("\xa0", " ").strip()

def get_known_names():
    global KNOWN_NAMES
    if KNOWN_NAMES is not None:
        return KNOWN_NAMES
    KNOWN_NAMES = set()
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT player_name FROM bwf_historical_rankings WHERE player_name IS NOT NULL")
        for row in cursor.fetchall():
            n = clean_text(row[0])
            if n:
                KNOWN_NAMES.add(n.lower())
        conn.close()
    except Exception as e:
        print(f"Warning: Could not load known names from DB: {e}")
    return KNOWN_NAMES

def detect_columns(header_line):
    header_line = clean_text(header_line).lower()
    tokens = header_line.split()
    
    is_doubles = any("p1" in t or "-1" in t or "country-1" in t or "name-1" in t or "id-1" in t or "country-2" in t or "name-2" in t or "id-2" in t for t in tokens)
    
    has_gender = "gender" in header_line or "gender" in "".join(tokens) or "gende" in header_line
    has_last_first = "last name" in header_line or "first name" in header_line or "last" in tokens or "first" in tokens
    
    layout = "unknown"
    if is_doubles:
        if "country-1" in header_line:
            layout = "doubles_2025"
        elif "p1 gender" in header_line or "p1gender" in "".join(tokens) or "p1 last" in header_line:
            layout = "doubles_2023"
        else:
            layout = "doubles_2026"
    else:
        if "player's name" in header_line or "player name" in header_line:
            layout = "singles_2025"
        elif "gender" in header_line or "gender" in "".join(tokens) or "last name" in header_line:
            layout = "singles_2023"
        else:
            layout = "singles_2026"
            
    return {
        "is_doubles": is_doubles,
        "layout": layout,
        "has_gender": has_gender,
        "has_last_first": has_last_first,
        "raw_header": header_line
    }

def split_doubles_names(combined_names, id1, id2, c1, c2):
    known = get_known_names()
    words = combined_names.split()
    n = len(words)
    if n <= 1:
        return combined_names, ""
        
    best_k = None
    best_score = -999999
    
    for k in range(1, n):
        p1_cand = " ".join(words[:k])
        p2_cand = " ".join(words[k:])
        
        p1_known = p1_cand.lower() in known
        p2_known = p2_cand.lower() in known
        
        score = 0
        if p1_known:
            score += 1000
        if p2_known:
            score += 1000
            
        p1_upper_cnt = sum(1 for w in words[:k] if w.isupper() and len(w) > 1)
        p2_upper_cnt = sum(1 for w in words[k:] if w.isupper() and len(w) > 1)
        
        if p1_upper_cnt > 0:
            score += 10
        if p2_upper_cnt > 0:
            score += 10
            
        if words[k-1].isupper() and len(words[k-1]) > 1:
            score += 5
        if words[0].isupper() and len(words[0]) > 1:
            score += 5
            
        if words[k].isupper() and len(words[k]) > 1:
            score += 5
        if words[-1].isupper() and len(words[-1]) > 1:
            score += 5
            
        score -= abs(k - (n - k)) * 2
        
        if score > best_score:
            best_score = score
            best_k = k
            
    p1 = " ".join(words[:best_k])
    p2 = " ".join(words[best_k:])
    return p1, p2
def normalize_tokens(tokens):
    if len(tokens) < 3:
        return tokens
        
    # 1. Merge adjacent single digits at the end of the list (before the last token)
    # E.g. ['5', '5', '1'] -> ['55', '1']
    i = len(tokens) - 2
    while i >= 0:
        if tokens[i].isdigit() and len(tokens[i]) == 1:
            j = i
            while j >= 0 and tokens[j].isdigit() and len(tokens[j]) == 1:
                j -= 1
            if j + 1 < i:
                merged_num = "".join(tokens[j+1:i+1])
                tokens[j+1:i+1] = [merged_num]
                i = j
            else:
                i -= 1
        else:
            i -= 1

    # 2. Merge adjacent uppercase tokens before the Points token
    numeric_indices = []
    for idx in range(len(tokens) - 1, -1, -1):
        if tokens[idx].isdigit():
            numeric_indices.append(idx)
            
    if len(numeric_indices) >= 2:
        pts_idx = numeric_indices[1]
    elif len(numeric_indices) == 1:
        pts_idx = numeric_indices[0]
    else:
        pts_idx = -1
        
    if pts_idx != -1:
        j = pts_idx - 1
        upper_seq = []
        # Limit to 4 tokens maximum to avoid scanning into the player name
        while j >= 0 and len(upper_seq) < 4:
            t = tokens[j]
            if t.isupper() and len(t) <= 2:
                upper_seq.insert(0, (j, t))
                j -= 1
            else:
                break
                
        if upper_seq:
            start_idx = upper_seq[0][0]
            end_idx = upper_seq[-1][0]
            merged = "".join([x[1] for x in upper_seq])
            
            if len(merged) == 4 and merged[0] in ['M', 'F']:
                replacement = [merged[0], merged[1:]]
            elif len(merged) == 3:
                replacement = [merged]
            else:
                replacement = list(tokens[start_idx:end_idx+1])
                
            tokens[start_idx:end_idx+1] = replacement
            
    # 3. Clean up name split: if the token before gender/country is 'f' (lowercase, len 1),
    # merge it with the token before it
    for idx, t in enumerate(tokens):
        if t in ['M', 'F'] or (len(t) == 3 and t.isupper() and not t.isdigit()):
            if idx > 0 and tokens[idx-1] == 'f':
                if idx > 1:
                    tokens[idx-2] = tokens[idx-2] + 'f'
                    tokens.pop(idx-1)
            break
            
    return tokens

def parse_line(line, schema):
    tokens = normalize_tokens(line.split())
    if not tokens:
        return None
        
    rank_str = tokens[0]
    rank_str = re.sub(r"\D", "", rank_str)
    if not rank_str:
        return None
    rank = int(rank_str)
    
    layout = schema["layout"]
    
    try:
        if layout == "singles_2025":
            country = tokens[1]
            if len(country) != 3 or not country.isupper():
                for idx, t in enumerate(tokens[1:4], start=1):
                    if len(t) == 3 and t.isupper():
                        country = t
                        break
            
            num_tokens = []
            for t in reversed(tokens):
                t_clean = re.sub(r"\D", "", t)
                if t_clean.isdigit():
                    num_tokens.append(int(t_clean))
                else:
                    break
            
            if len(num_tokens) >= 3:
                tour = num_tokens[0]
                points = num_tokens[1]
                player_id = num_tokens[2]
                name_tokens = tokens[2:-3]
            elif len(num_tokens) == 2:
                tour = 0
                points = num_tokens[0]
                player_id = num_tokens[1]
                name_tokens = tokens[2:-2]
            else:
                return None
                
            player_name = " ".join(name_tokens)
            return {"rank": rank, "player_id": player_id, "player_name": player_name, "country": country, "points": points}
            
        elif layout == "singles_2026":
            player_id = int(re.sub(r"\D", "", tokens[1]))
            
            num_tokens = []
            country = ""
            name_end_idx = -1
            for idx, t in enumerate(reversed(tokens)):
                t_clean = re.sub(r"\D", "", t)
                if t_clean.isdigit():
                    num_tokens.append(int(t_clean))
                elif len(t) == 3 and t.isupper():
                    country = t
                    name_end_idx = len(tokens) - idx - 1
                    break
            else:
                return None
                
            if len(num_tokens) >= 2:
                tour = num_tokens[0]
                points = num_tokens[1]
            elif len(num_tokens) == 1:
                tour = 0
                points = num_tokens[0]
            else:
                return None
                
            name_tokens = tokens[2:name_end_idx]
            player_name = " ".join(name_tokens)
            return {"rank": rank, "player_id": player_id, "player_name": player_name, "country": country, "points": points}
            
        elif layout == "singles_2023":
            player_id = int(re.sub(r"\D", "", tokens[1]))
            
            num_tokens = []
            country = ""
            gender_idx = -1
            for idx, t in enumerate(reversed(tokens)):
                t_clean = re.sub(r"\D", "", t)
                if t_clean.isdigit():
                    num_tokens.append(int(t_clean))
                elif len(t) == 3 and t.isupper() and not country:
                    country = t
                elif t in ["M", "F"]:
                    gender_idx = len(tokens) - idx - 1
                    break
            else:
                for idx, t in enumerate(reversed(tokens)):
                    if len(t) == 3 and t.isupper() and not t.isdigit():
                        country = t
                        gender_idx = len(tokens) - idx - 1
                        break
                        
            if len(num_tokens) >= 2:
                tour = num_tokens[0]
                points = num_tokens[1]
            elif len(num_tokens) == 1:
                tour = 0
                points = num_tokens[0]
            else:
                return None
                
            name_tokens = tokens[2:gender_idx]
            player_name = " ".join(name_tokens)
            return {"rank": rank, "player_id": player_id, "player_name": player_name, "country": country, "points": points}
            
        elif layout == "doubles_2025":
            c1 = tokens[1]
            c2 = tokens[2]
            
            num_tokens = []
            for t in reversed(tokens):
                t_clean = re.sub(r"\D", "", t)
                if t_clean.isdigit():
                    num_tokens.append(int(t_clean))
                else:
                    break
            
            if len(num_tokens) >= 4:
                tour = num_tokens[0]
                points = num_tokens[1]
                id2 = num_tokens[2]
                id1 = num_tokens[3]
                name_tokens = tokens[3:-4]
            elif len(num_tokens) == 3:
                tour = 0
                points = num_tokens[0]
                id2 = num_tokens[1]
                id1 = num_tokens[2]
                name_tokens = tokens[3:-3]
            else:
                return None
                
            combined_names = " ".join(name_tokens)
            p1_name, p2_name = split_doubles_names(combined_names, id1, id2, c1, c2)
            
            return {
                "rank": rank,
                "p1_id": id1, "p1_name": p1_name, "p1_country": c1,
                "p2_id": id2, "p2_name": p2_name, "p2_country": c2,
                "points": points
            }
            
        elif layout == "doubles_2026":
            id1 = int(re.sub(r"\D", "", tokens[1]))
            
            id2_idx = -1
            for i in range(2, len(tokens)):
                if tokens[i].isdigit():
                    id2_idx = i
                    break
            
            if id2_idx == -1:
                return None
                
            id2 = int(tokens[id2_idx])
            
            c1 = tokens[id2_idx - 1]
            p1_name = " ".join(tokens[2:id2_idx - 1])
            
            num_tokens = []
            c2 = ""
            c2_idx = -1
            for idx, t in enumerate(reversed(tokens)):
                t_clean = re.sub(r"\D", "", t)
                if t_clean.isdigit():
                    num_tokens.append(int(t_clean))
                elif len(t) == 3 and t.isupper() and not c2:
                    c2 = t
                    c2_idx = len(tokens) - idx - 1
                    break
            
            if c2_idx == -1 or len(num_tokens) < 1:
                return None
                
            points = num_tokens[-1]
            p2_name = " ".join(tokens[id2_idx + 1:c2_idx])
            
            return {
                "rank": rank,
                "p1_id": id1, "p1_name": p1_name, "p1_country": c1,
                "p2_id": id2, "p2_name": p2_name, "p2_country": c2,
                "points": points
            }
            
        elif layout == "doubles_2023":
            id1 = int(re.sub(r"\D", "", tokens[1]))
            
            id2_idx = -1
            for i in range(2, len(tokens)):
                if tokens[i].isdigit():
                    id2_idx = i
                    break
            
            if id2_idx == -1:
                return None
                
            id2 = int(tokens[id2_idx])
            
            c1 = tokens[id2_idx - 1]
            p1_end = id2_idx - 1
            if tokens[id2_idx - 2] in ["M", "F"]:
                p1_end = id2_idx - 2
            p1_name = " ".join(tokens[2:p1_end])
            
            num_tokens = []
            c2 = ""
            c2_idx = -1
            for idx, t in enumerate(reversed(tokens)):
                t_clean = re.sub(r"\D", "", t)
                if t_clean.isdigit():
                    num_tokens.append(int(t_clean))
                elif len(t) == 3 and t.isupper() and not c2:
                    c2 = t
                    c2_idx = len(tokens) - idx - 1
                    break
            
            if c2_idx == -1 or len(num_tokens) < 1:
                return None
                
            points = num_tokens[-1]
            
            p2_end = c2_idx
            if tokens[c2_idx - 1] in ["M", "F"]:
                p2_end = c2_idx - 1
            p2_name = " ".join(tokens[id2_idx + 1:p2_end])
            
            return {
                "rank": rank,
                "p1_id": id1, "p1_name": p1_name, "p1_country": c1,
                "p2_id": id2, "p2_name": p2_name, "p2_country": c2,
                "points": points
            }
            
    except Exception as e:
        return None
        
    return None

def parse_pdf(file_path, rank_date, week):
    print(f"Parsing PDF: {os.path.basename(file_path)}")
    rows_to_insert = []
    
    try:
        reader = pypdf.PdfReader(file_path)
    except Exception as e:
        print(f"  [!] Error opening PDF file: {e}")
        return []
        
    current_event = None
    current_schema = None
    row_count = 0
    
    for page_idx, page in enumerate(reader.pages):
        text = page.extract_text()
        text_clean = clean_text(text)
        
        # Detect event changes on pages
        text_upper = text_clean.upper()
        if "MEN'S SINGLES" in text_upper and "WOMEN'S SINGLES" not in text_upper:
            current_event = "MS"
        elif "WOMEN'S SINGLES" in text_upper:
            current_event = "WS"
        elif "MEN'S DOUBLES" in text_upper and "WOMEN'S DOUBLES" not in text_upper:
            current_event = "MD"
        elif "WOMEN'S DOUBLES" in text_upper:
            current_event = "WD"
        elif "MIXED DOUBLES" in text_upper:
            current_event = "XD"
            
        raw_lines = text.split("\n")
        lines = []
        for rl in raw_lines:
            rl_clean = clean_text(rl)
            if not rl_clean:
                continue
            tokens = rl_clean.split()
            first_token = tokens[0]
            first_token_digits = re.sub(r"\D", "", first_token)
            if first_token_digits and lines:
                lines.append(rl_clean)
            else:
                if lines:
                    lines[-1] = lines[-1] + " " + rl_clean
                else:
                    lines.append(rl_clean)
        
        header_line = None
        for line in lines:
            line_clean = clean_text(line).lower()
            if "rank" in line_clean and ("points" in line_clean or "point" in line_clean or "bwf id" in line_clean):
                header_line = line
                break
                
        if header_line:
            detected = detect_columns(header_line)
            if detected["is_doubles"]:
                if current_event not in ["MD", "WD", "XD"]:
                    current_event = "MD"
            else:
                if current_event not in ["MS", "WS"]:
                    current_event = "MS"
            current_schema = detected
            if current_event in ["MD", "WD", "XD"] and not current_schema["is_doubles"]:
                current_schema["is_doubles"] = True
                if "doubles" not in current_schema["layout"]:
                    current_schema["layout"] = current_schema["layout"].replace("singles", "doubles")
            elif current_event in ["MS", "WS"] and current_schema["is_doubles"]:
                current_schema["is_doubles"] = False
                if "singles" not in current_schema["layout"]:
                    current_schema["layout"] = current_schema["layout"].replace("doubles", "singles")
        
        if not current_schema or not current_event:
            continue
            
        for line in lines:
            line_clean = clean_text(line)
            if not line_clean:
                continue
                
            if "rank" in line_clean.lower() and ("points" in line_clean.lower() or "point" in line_clean.lower()):
                continue
                
            parsed = parse_line(line_clean, current_schema)
            if parsed:
                if current_schema["is_doubles"]:
                    rows_to_insert.append((rank_date, week, current_event, parsed["rank"], parsed["p1_id"], parsed["p1_name"], parsed["p1_country"], parsed["points"]))
                    rows_to_insert.append((rank_date, week, current_event, parsed["rank"], parsed["p2_id"], parsed["p2_name"], parsed["p2_country"], parsed["points"]))
                    row_count += 2
                else:
                    rows_to_insert.append((rank_date, week, current_event, parsed["rank"], parsed["player_id"], parsed["player_name"], parsed["country"], parsed["points"]))
                    row_count += 1
                    
    print(f"  [OK] Parsed {row_count} total rows across all disciplines.")
    return rows_to_insert

def main():
    if not os.path.exists(INPUT_DIR):
        print(f"No downloads directory found at {INPUT_DIR}.")
        return
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Initialize schema first to be safe
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bwf_historical_rankings (
            rank_date TEXT NOT NULL,
            week INTEGER NOT NULL,
            event TEXT NOT NULL,
            rank INTEGER NOT NULL,
            player_id INTEGER NOT NULL,
            player_name TEXT,
            country TEXT,
            points INTEGER,
            PRIMARY KEY (rank_date, event, rank, player_id)
        );
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_bwf_hist_player ON bwf_historical_rankings(player_id, rank_date)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_bwf_hist_date ON bwf_historical_rankings(rank_date, event)")
    conn.commit()
    
    # Get set of already parsed dates
    cursor.execute("SELECT DISTINCT rank_date FROM bwf_historical_rankings")
    parsed_dates = {row[0] for row in cursor.fetchall()}
    print(f"Found {len(parsed_dates)} dates already loaded in database.")
    
    files = [f for f in os.listdir(INPUT_DIR) if f.startswith("WR_") and f.lower().endswith((".xlsx", ".pdf"))]
    print(f"Found {len(files)} files to parse in {INPUT_DIR}.")
    
    parsed_count = 0
    for filename in sorted(files):
        rank_date, week = parse_date_week_from_filename(filename)
        if not rank_date or week is None:
            continue
            
        if rank_date in parsed_dates:
            # Already parsed, skip
            continue
            
        file_path = os.path.join(INPUT_DIR, filename)
        if os.path.getsize(file_path) == 0:
            print(f"Skipping empty or incomplete file: {filename}")
            continue
            
        # Parse based on file type
        if filename.lower().endswith(".xlsx"):
            rows = parse_xlsx(file_path, rank_date, week)
        else:
            rows = parse_pdf(file_path, rank_date, week)
            
        if rows:
            cursor.executemany("""
                INSERT OR REPLACE INTO bwf_historical_rankings 
                (rank_date, week, event, rank, player_id, player_name, country, points)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, rows)
            conn.commit()
            print(f"  [SUCCESS] Ingested {len(rows)} entries for {rank_date} (Week {week}) into database.")
            parsed_count += 1
            
    conn.close()
    print(f"\n==========================================")
    print(f" PARSER COMPLETE")
    print(f"==========================================")
    print(f"Newly Ingested Weeks : {parsed_count}")
    print("==========================================")

if __name__ == "__main__":
    main()
